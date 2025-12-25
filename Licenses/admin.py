from django.contrib import admin
from .models import License, Arrival
from Issuance.models import Logbook
from django.contrib import admin, messages
from django.shortcuts import render, redirect
from django.urls import path
from django.db import transaction
import openpyxl
from datetime import datetime
from .models import License, Arrival
from .forms import UploadExcelForm

class ArrivalInline(admin.TabularInline):
    save_on_top = True
    model = Arrival
    extra = 0
    # readonly_fields = ('date', 'request_number', 'network_number', 'organization_name', 'inn', 'valid_until', 'license_object', 'quantity', 'uploaded', 'uploaded_date', 'note')
    readonly_fields = ()
    fields = ('date', 'request_number', 'network_number', 'organization_name', 'inn', 'valid_until', 'license_object', 'quantity', 'uploaded', 'uploaded_date', 'note')
    # При сохранении из inline admin будет вызван save() модели Arrival

class IssuanceInline(admin.TabularInline):
    save_on_top = True
    model = Logbook
    extra = 0
    readonly_fields = ('net_number', 'log_number', 'number_naumen', 'number_elk', 'ogv', 'amount', 'platform')
    fields = ('net_number', 'log_number', 'number_naumen', 'number_elk', 'ogv', 'amount', 'platform')

@admin.register(License)
class LicenseAdmin(admin.ModelAdmin):
    list_display = ('network_number', 'organization_name', 'inn', 'license_object', 'total_received', 'used', 'free', 'last_replenishment')
    list_display_links = ('organization_name',)
    search_fields = ('organization_name', 'inn', 'license_object', 'network_number', 'total_received')
    list_filter = ( 'network_number', 'last_replenishment', 'free', 'license_object')
    inlines = [ArrivalInline, IssuanceInline]
    save_on_top = True


    change_list_template = 'admin/licenses/license_change_list.html'  # шаблон со кнопкой импорта

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-arrivals/', self.admin_site.admin_view(self.import_arrivals), name='licenses_import_arrivals'),
        ]
        return my_urls + urls

    def import_arrivals(self, request):
        """
        View для загрузки .xlsx и создания Arrival-объектов.
        Ожидается таблица с заголовком в первой строке.
        Поддерживаем заголовки на русском и английском: date/дата, request_number/номер заявки, network_number/№ сети, organization_name/наименование организации,
        inn/инн, valid_until/срок действия лицензии, license_object/объект лицензии, quantity/количество, uploaded/загружено, uploaded_date/дата загрузки, note/примечание
        """
        if request.method == 'POST':
            form = UploadExcelForm(request.POST, request.FILES)
            if form.is_valid():
                uploaded_file = form.cleaned_data['file']
                try:
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                except Exception as e:
                    messages.error(request, f'Не удалось открыть файл Excel: {e}')
                    return redirect('..')

                ws = wb.active

                # Считаем заголовки (первый ряд)
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = []
                for h in header_row:
                    if h is None:
                        headers.append('')
                    else:
                        headers.append(str(h).strip().lower())

                # Словарь соответствия возможных заголовков -> поля Arrival
                header_map = {
                    'date': 'date',
                    'дата': 'date',
                    'request_number': 'request_number',
                    'номер заявки': 'request_number',
                    'request number': 'request_number',
                    'network_number': 'network_number',
                    'номер сети': 'network_number',
                    'network': 'network_number',
                    'organization_name': 'organization_name',
                    'организация': 'organization_name',
                    'наименование организации': 'organization_name',
                    'inn': 'inn',
                    'инн': 'inn',
                    'valid_until': 'valid_until',
                    'срок действия лицензии': 'valid_until',
                    'license_object': 'license_object',
                    'объект лицензии': 'license_object',
                    'quantity': 'quantity',
                    'количество': 'quantity',
                    'uploaded': 'uploaded',
                    'загружено': 'uploaded',
                    'uploaded_date': 'uploaded_date',
                    'дата загрузки': 'uploaded_date',
                    'note': 'note',
                    'примечание': 'note',
                }

                # map column index -> field
                mapping = {}
                for idx, h in enumerate(headers):
                    if h in header_map:
                        mapping[header_map[h]] = idx

                required = ['date','request_number','network_number','organization_name','inn','license_object','quantity']
                missing = [f for f in required if f not in mapping]
                if missing:
                    messages.error(request, f'В файле отсутствуют обязательные столбцы: {missing}')
                    return redirect('..')

                success_count = 0
                errors = []

                # Импорт в одной транзакции: при любом исключении — откат всего импорта
                try:
                    with transaction.atomic():
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            try:
                                def get_cell(field):
                                    idx = mapping.get(field)
                                    return row[idx] if (idx is not None and idx < len(row)) else None

                                # Преобразования значений
                                date_val = get_cell('date')
                                if isinstance(date_val, str):
                                    try:
                                        date_val = datetime.strptime(date_val.strip(), '%d.%m.%Y').date()
                                    except Exception:
                                        try:
                                            date_val = datetime.fromisoformat(date_val.strip()).date()
                                        except Exception:
                                            date_val = None

                                request_number = get_cell('request_number') or ''
                                network_number = get_cell('network_number')
                                if network_number is None or str(network_number).strip()=='':
                                    network_number = 0
                                else:
                                    network_number = int(network_number)

                                organization_name = get_cell('organization_name') or ''
                                inn = get_cell('inn')
                                if inn is None or str(inn).strip()=='':
                                    inn = None
                                else:
                                    inn = int(inn)

                                valid_until = get_cell('valid_until')
                                if isinstance(valid_until, str):
                                    try:
                                        valid_until = datetime.strptime(valid_until.strip(), '%d.%m.%Y').date()
                                    except Exception:
                                        try:
                                            valid_until = datetime.fromisoformat(valid_until.strip()).date()
                                        except Exception:
                                            valid_until = None

                                license_object = get_cell('license_object') or ''
                                quantity = get_cell('quantity') or 0
                                quantity = int(quantity)

                                uploaded = get_cell('uploaded')
                                if isinstance(uploaded, str):
                                    uploaded = uploaded.strip().lower() in ('1','true','да','yes','y')
                                else:
                                    uploaded = bool(uploaded)

                                uploaded_date = get_cell('uploaded_date')
                                if isinstance(uploaded_date, str):
                                    try:
                                        uploaded_date = datetime.strptime(uploaded_date.strip(), '%d.%m.%Y').date()
                                    except Exception:
                                        try:
                                            uploaded_date = datetime.fromisoformat(uploaded_date.strip()).date()
                                        except Exception:
                                            uploaded_date = None

                                note = get_cell('note') or ''

                                arrival = Arrival(
                                    date=date_val,
                                    request_number=str(request_number),
                                    network_number=network_number,
                                    organization_name=str(organization_name),
                                    inn=inn,
                                    valid_until=valid_until,
                                    license_object=str(license_object),
                                    quantity=quantity,
                                    uploaded=uploaded,
                                    uploaded_date=uploaded_date,
                                    note=str(note),
                                )
                                # валидируем по clean()
                                arrival.full_clean()
                                # save() выполнит всю логику обновления License
                                arrival.save()
                                success_count += 1
                            except Exception as e:
                                errors.append(f'Строка {row_idx}: {e}')
                        # если есть ошибки — выбрасываем исключение для отката
                        if errors:
                            raise Exception('\\n'.join(errors))
                except Exception as e:
                    messages.error(request, f'Ошибки при импорте: {e}')
                    return redirect('..')

                messages.success(request, f'Импорт завершён. Добавлено записей: {success_count}')
                return redirect('..')
        else:
            form = UploadExcelForm()

        context = dict(self.admin_site.each_context(request))
        context.update({
            'form': form,
            'title': 'Импорт поступлений из Excel',
        })
        return render(request, 'admin/licenses/import_arrivals.html', context)

@admin.register(Arrival)
class ArrivalAdmin(admin.ModelAdmin):
    list_display = ('request_number', 'date', 'organization_name', 'inn', 'license_object', 'quantity', 'uploaded', 'uploaded_date')
    search_fields = ('request_number', 'organization_name', 'inn', 'license_object')
    list_filter = ('uploaded', 'date')

    def save_model(self, request, obj, form, change):
        obj.full_clean()
        super().save_model(request, obj, form, change)



